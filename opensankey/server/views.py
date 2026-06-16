"""
==================================================================================================
The MIT License (MIT)
==================================================================================================
Copyright (c) 2025 TerriFlux

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in
all copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
THE SOFTWARE.
==================================================================================================
Author        : Vincent LE DOZE & Vincent CLAVEL & Julien Alapetite for TerriFlux
==================================================================================================
"""

# coding: utf-8
# flake8: noqa

from pathlib import Path
import tempfile
import os
import json
import shutil
from time import perf_counter

import pandas as pd
from .views_utils import cut_layout
import openpyxl

from .views_utils import clean_file, handle_json_or_compressed, parse_folder
import requests

from threading import Thread, Lock

# Flask modules imports
from flask import abort, make_response
from flask import Blueprint
from flask import current_app
from flask import render_template
from flask import request
from flask import Response
from flask import send_file
from flask import send_from_directory
from flask import session

import SankeyExcelParser.su_trace as trace
from SankeyExcelParser.io_base import IOExcel, IOJson
from . import sankeymatic

template_folder = os.path.join(
    os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "client"),
    "build",
)

static_folder = os.path.join(
    os.path.join(
        os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "client"),
        "build",
    ),
    "static",
)

opensankey = Blueprint(
    "opensankey",
    __name__,
    static_folder=static_folder,
    template_folder=template_folder,
    static_url_path="/static/opensankey",
)

image_template_folder = os.path.join(
    os.path.join(
        os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "client"),
        "src",
    ),
    "images",
)


def get_process_state():  # ← Plus de paramètre session_id
    """Récupère l'état depuis Flask session"""
    return session.get('process_state', {
        'process_started': False,
        'logname': None,
        'output_file_name': None,
        'input_format': None,
        'output_format': None
    })


def set_process_state(**kwargs):  # ← Plus de paramètre session_id
    """Stocke l'état dans Flask session"""
    if 'process_state' not in session:
        session['process_state'] = {}
    session['process_state'].update(kwargs)
    session.modified = True  # ← IMPORTANT !


# --- Statut de traitement piloté par fichier (canal cross-thread) -----------
# Les traitements lourds (conversion, optimisation) tournent dans un Thread
# détaché qui n'a PAS de contexte requête Flask : il ne peut donc pas écrire
# dans la session (cf. set_process_state). Le seul canal partagé fiable depuis
# ce thread est le système de fichiers — déjà utilisé pour le log. On écrit un
# petit fichier de statut frère du log (<logname>.status) que le thread met à
# jour à chaque sortie, et que check_process (en contexte requête) relit pour
# piloter l'arrêt du polling côté client. Cela remplace le grep de prose
# localisée (FINISHED/TERMINÉ/ÉCHOUÉ…) qui était fragile et couplé à la langue.
PROCESS_STATUS_RUNNING = "running"
PROCESS_STATUS_FINISHED = "finished"
PROCESS_STATUS_FAILED = "failed"


def _process_status_path(log_filename):
    """Chemin du fichier de statut associé à un fichier log."""
    if not log_filename:
        return None
    return log_filename + ".status"


def write_process_status(log_filename, status):
    """Écrit le statut de traitement. Appelable depuis le thread (best-effort)."""
    path = _process_status_path(log_filename)
    if not path:
        return
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write(status)
    except OSError:
        # Canal best-effort : en cas d'échec d'écriture, le client retombe sur
        # l'absence de statut (le polling continue) plutôt que de casser le run.
        pass


def read_process_status(log_filename):
    """Relit le statut écrit par le thread. None si absent/illisible (= en cours)."""
    path = _process_status_path(log_filename)
    if not path or not os.path.isfile(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return f.read().strip() or None
    except OSError:
        return None


@opensankey.route("/")
def start():
    return render_template("index.html", filename="", static_site="false")


@opensankey.route("/<adress>")
def goto(adress):
    return render_template(adress)


@opensankey.route("/upload/check_process", methods=["POST"])
def check_process():
    state = get_process_state()
    if not state['process_started']:
        if 'logname' not in state:
            return Response(json.dumps({}), status=200, mimetype="application/json")
        return Response(json.dumps({"not_started": True}), status=200, mimetype="application/json")
    try:
        logname = state['logname']
        if os.path.isfile(logname):
            # Le log est écrit en UTF-8 (su_trace.logger_init :
            # logging.FileHandler(..., encoding="utf-8")). Sous Windows, open()
            # sans encodage explicite décode en cp1252 et lève UnicodeDecodeError
            # dès qu'un nom de feuille/nœud accentué produit un octet hors cp1252
            # → 500 en boucle, spinner client figé. On lit donc en UTF-8 tolérant.
            with open(logname, "r", encoding="utf-8", errors="replace") as f:
                results = f.read()
            results_dict = {
                "log_name": logname,
                "output": results,
                # Statut machine-lisible piloté par le thread via <logname>.status.
                # None tant que le thread n'a rien écrit (= traitement en cours) ;
                # le client s'arrête sur 'finished'/'failed', plus sur le texte.
                "status": read_process_status(logname),
            }
            json_data = json.dumps(results_dict)
            # trace.logger.debug('dumps')
            return Response(json_data, status=200, mimetype="application/json")
        else:
            return Response(
                json.dumps({"output": "ERROR: /upload/check_process: le fichier tmp_log n'existe pas."}),
                status=500,
                mimetype="application/json",
            )
    except json.JSONDecodeError:
        return Response(
            json.dumps({"output": "ERROR: /upload/check_process: le fichier tmp_log ne peut pas être ouvert."}),
            status=500,
            mimetype="application/json",
        )


@opensankey.route("/upload/retrieve_result", methods=["POST"])
def retrieve_result():
    """
    Route générique pour récupérer le fichier résultat d'un traitement.
    Gère automatiquement le mimetype selon l'extension du fichier.
    Fonctionne pour :
    - Conversions (Excel, Pickle, JSON)
    - Chargements Excel/Pickle → JSON
    """
    state = get_process_state()

    # Marquer le processus comme terminé
    set_process_state(process_started=False)

    try:
        output_file_name = state.get("output_file_name")
        # Debug mode (reconciliation): SA pre-stashes a path to a .zip bundle
        # produced by solve_optimisation_problem_unified once the run finishes.
        # If that zip materialised, hand it back; otherwise fall through to the
        # plain output (covers cases where MFAProblem couldn't write the
        # constraints_summary.txt and the zip never got assembled).
        debug_zip_path = state.get("debug_zip_path")
        if debug_zip_path and os.path.exists(debug_zip_path):
            output_file_name = debug_zip_path
        json_gz_path = output_file_name
        if output_file_name.endswith(".json"):
            json_gz_path = output_file_name + ".gz"

        if not output_file_name:
            return Response(
                json.dumps({"error": "No active session"}),
                status=400,
                mimetype="application/json"
            )

        if not os.path.exists(output_file_name) and not os.path.exists(json_gz_path):
            return Response(
                json.dumps({"error": "Output file not found"}),
                status=404,
                mimetype="application/json"
            )
        # Déterminer l'extension et le mimetype
        root_file_name, ext = os.path.splitext(output_file_name)
        ext = ext.lower()
        output_format = request.form.get("output_format", "json")
        if output_format == 'excel' and ext == '.json':
            io_json = IOJson()
            ok, msg = io_json.load_sankey(output_file_name)
            if not ok:
                trace.logger.error(f"-- ERROR loading sankey from Pickle: {msg}")
                trace.logger.info("{:-<{w}}".format(" [FAILED] Could not load sankey", w=120))
                return Response(
                    json.dumps({"error": "Output file not found"}),
                    status=405,
                    mimetype="application/json"
                )
            io_excel = IOExcel(io_json.sankey)
            io_excel.write_sankey(root_file_name + ".xlsx")
            ext = ".xlsx"
            output_file_name = root_file_name + ".xlsx"

        # Mapping extensions → mimetypes
        mimetype_map = {
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            '.pkl': 'application/octet-stream',
            '.json': 'application/json',
            '.csv': 'text/csv',
            '.txt': 'text/plain',
            '.zip': 'application/zip'
        }

        # Mapping extensions → noms de téléchargement
        download_name_map = {
            '.xlsx': root_file_name + '.xlsx',
            '.json': root_file_name + '.json',
            '.csv': root_file_name + '.csv',
            '.txt': root_file_name + '.txt',
            '.zip': root_file_name + '.zip'
        }

        mimetype = mimetype_map.get(ext, 'application/octet-stream')
        download_name = download_name_map.get(ext, root_file_name + ext)

        trace.logger.info(f"Envoi du fichier: {output_file_name} ({mimetype})")
        # Skip the json gzip step when the actual file is already a zip bundle
        # (debug-mode reconciliation reroutes the output to a zip containing
        # both the reconciled file and constraints_summary.txt).
        if output_format == 'json' and ext != '.zip':
            output_file_name = handle_json_or_compressed(output_file_name)
        return send_file(
            output_file_name,
            as_attachment=True,
            download_name=download_name,
            mimetype=mimetype
        )

    except Exception as excpt:
        trace.logger.error(f"retrieve_result failed: {str(excpt)}")
        return Response(
            json.dumps({"error": str(excpt)}),
            status=500,
            mimetype="application/json"
        )


@opensankey.route("/upload/retrieve_json", methods=["POST"])
def retrieve_json():
    """
    Route générique pour récupérer le fichier résultat d'un traitement.
    """
    state = get_process_state()

    log_dir = tempfile.mkdtemp()
    log_filename = log_dir + os.path.sep + "rollover.log"
    trace.logger_init(log_filename, "w")

    set_process_state(
        process_started=True,
        logname=log_filename
    )

    try:
        input_file_name = state.get("output_file_name")

        if not input_file_name:
            return Response(
                json.dumps({"error": "No active session"}),
                status=400,
                mimetype="application/json"
            )

        if not os.path.exists(input_file_name):
            return Response(
                json.dumps({"error": "Output file not found"}),
                status=404,
                mimetype="application/json"
            )

        output_file_name = os.path.join(os.path.dirname(input_file_name), "output.json")

        input_format = state.get('input_format', 'excel')
        logname = state.get('logname')

        # Mettre à jour l'état
        set_process_state(
            input_filename=input_file_name,
            output_file_name=output_file_name,
            input_format=input_format,
            output_format='json'
        )

        # Use threading for large files
        thread = Thread(
            target=conversion_thread,
            args=(
                input_file_name,
                output_file_name,
                input_format,
                'json',
                {},  # input_options
                {},  # output_options
                logname,
                {}   # sankey_as_data
            ),
        )
        thread.daemon = True
        thread.start()
        trace.logger.debug("Conversion thread launched")

        return Response(response="{}", status=200, mimetype="application/json")

    except Exception as excpt:
        trace.logger.error(f"launch_conversion failed: {str(excpt)}")
        return Response(
            json.dumps({"error": str(excpt)}),
            status=500,
            mimetype="application/json"
        )


@opensankey.route("/convert/peek_options", methods=["POST"])
def peek_options():
    """
    [IID=162] Lightweight, synchronous peek of an uploaded Excel workbook's
    "Options de réconciliation" sheet.

    Returns the key-value options it declares (input autocorrection flags +
    solver flags) so the converter / reconciliation dialog can pre-tick the
    matching checkboxes the moment the file is selected — before the full parse
    runs, which is otherwise too late to inform the load options. Reuses the
    parser's own sheet-name recognition (accent-insensitive) and value coercion
    so it stays consistent with an actual load. Always answers 200 with a
    (possibly empty) ``solver_options`` map; never blocks file selection.
    """
    from SankeyExcelParser.classes.sankey_pandas import SankeyPandas
    import SankeyExcelParser.io_excel_constants as PEEK_CONST
    options = {}
    tmp_dir = None
    try:
        uploaded = request.files.get("file")
        if uploaded is not None:
            tmp_dir = tempfile.mkdtemp()
            path = os.path.join(tmp_dir, "peek.xlsx")
            uploaded.save(path)
            sankey = SankeyPandas()
            xl = pd.ExcelFile(path)
            for sheet_name in xl.sheet_names:
                ok, refkey = sankey._consistantSheetName(sheet_name)
                if not ok:
                    continue
                if refkey == PEEK_CONST.MFA_OPTIONS_SHEET:
                    ok_read, _ = sankey.xl_read_mfa_options_sheet(xl.parse(sheet_name))
                    if ok_read:
                        options = dict(sankey.mfa_options)
                elif refkey == PEEK_CONST.TAG_SHEET:
                    # #161 — parse the Tags sheet so the per dataTag-group
                    # propagate_structure flags are known (see derivation below).
                    try:
                        sankey.xl_read_tags_sheet(xl.parse(sheet_name))
                    except Exception:
                        pass
            # #161 — initialise the dialog's global "Propager la structure"
            # checkbox from the file: unchecked as soon as one dataTag group is
            # non-propagating. An explicit value in the Options sheet wins.
            data_groups = list(sankey.taggs.get(PEEK_CONST.TAG_TYPE_DATA, {}).values())
            if data_groups and "propagate_datatag_structure" not in options:
                options = dict(options)
                options["propagate_datatag_structure"] = all(
                    g.propagate_structure for g in data_groups
                )
    except Exception as e:
        trace.logger.warning(f"peek_options: {e}")
        options = {}
    finally:
        if tmp_dir is not None:
            shutil.rmtree(tmp_dir, ignore_errors=True)
    return Response(
        json.dumps({"solver_options": options}),
        status=200,
        mimetype="application/json",
    )


@opensankey.route("/convert/launch", methods=["POST"])
def launch_conversion():
    """
    Route universelle pour lancer une conversion de format.
    Paramètres attendus dans le formulaire :
    - file : fichier à convertir
    - input_format : 'excel' ou 'json'
    - output_format : 'excel' ou 'json'
    """
    try:
        tmp_dir = tempfile.mkdtemp()  # diff
        log_dir = tempfile.mkdtemp()
        log_filename = log_dir + os.path.sep + "rollover.log"
        # session["logname"] = log_filename
        trace.logger_init(log_filename, "w")

        input_format = request.form.get("input_format", "excel")
        output_format = request.form.get("output_format", "json")
        # Libellé localisé fourni par le dialogue appelant (Chargement / Édition /
        # Création d'index…) pour contextualiser le bandeau ; None => libellé
        # technique générique.
        process_label = request.form.get("process_label") or None

        # Keep input/output options separate: they share key names (e.g.
        # `activate_data_table`) but mean opposite things on each side.
        input_options = json.loads(request.form.get('input_options', '{}'))
        output_options = json.loads(request.form.get('output_options', '{}'))
        ext_map = {
            'excel': '.xlsx',
            'json': '.json',
            'blob': '.json'
        }
        if input_format != "example_excel" and input_format != "example_json":
            input_file_name = os.path.join(tmp_dir, f"input{ext_map[input_format]}")
        output_file_name = os.path.join(tmp_dir, f"output{ext_map[output_format]}")

        if input_format == "example_excel" or input_format == "example_json":
            exemple = request.form["file_name"]
            # Tout le contenu servi (templates/tutoriels) vit dans le submodule
            # SankeyData (env SANKEY_DATA, posee automatiquement par app.py depuis
            # la racine du checkout). Pas de repli MFAData. NB : on ne peut PAS
            # deduire le chemin depuis __file__ ici, car opensankey est installe
            # (copie) en site-packages cote serveur.
            sankey_data = os.environ.get("SANKEY_DATA")
            input_file_name = os.path.join(sankey_data, exemple) if sankey_data else exemple
            if input_format == "example_json":
                # Tolerance .json / .json.gz + conversion automatique en .json.gz
                # au premier chargement (mise en cache sur disque), exactement
                # comme pour les tutoriels.
                resolved = handle_json_or_compressed(input_file_name)
                if isinstance(resolved, str):
                    input_file_name = resolved
            extension = os.path.splitext(input_file_name)[1]
            if extension == '.xlsx':
                input_format = 'excel'
            else:
                # input_format = 'json'
                output_file_name = input_file_name
                set_process_state(
                    process_started=True,
                    input_filename=input_file_name,
                    output_file_name=output_file_name,
                    input_format=input_format,
                    output_format=output_format,
                    logname=log_filename
                )
                trace.logger.info("{:->{w}}".format(" " + (process_label or "CHARGEMENT TERMINÉE"), w=50))
                # Pas de thread ici (chargement direct d'un exemple JSON) : on
                # marque le statut terminé tout de suite pour arrêter le polling.
                write_process_status(log_filename, PROCESS_STATUS_FINISHED)
                return Response(response="{}", status=200, mimetype="application/json")
                # return handle_json_or_compressed(data_folder, exemple, input_file_name)

        elif input_format != 'example_excel' and input_format != 'example_json' and input_format != 'blob':
            input_file = request.files["file"]
            input_file.save(input_file_name)

        # Stocker l'état dans le stockage global
        set_process_state(
            process_started=True,
            input_filename=input_file_name,
            output_file_name=output_file_name,
            input_format=input_format,
            output_format=output_format,
            logname=log_filename
        )
        sankey_as_data = None
        if (input_format == "blob"):
            data = request.form["data"]
            sankey_as_data = data
            sankey_as_json = json.loads(sankey_as_data)
            io_json = IOJson()
            ok, log = io_json.load_sankey_from_json(sankey_as_json, do_coherence_checks=False)
            if not ok:
                trace.logger.error(f"FAILED load_sankey_from_json failed: {log}")
                return Response(
                    json.dumps({"error": str(log)}),
                    status=500,
                    mimetype="application/json"
                )
            io_json.write_sankey(input_file_name)
            input_format = "json"

        # Decide threading based on file size
        # file_stats = os.stat(input_file_name)
        # use_thread = file_stats.st_size > 500000  # 500KB threshold

        # if use_thread:
        # Use threading for large files
        thread = Thread(
            target=conversion_thread,
            args=(
                input_file_name,
                output_file_name,
                input_format,
                output_format,
                input_options,
                output_options,
                log_filename,
                sankey_as_data,
                process_label,
            ),
        )
        thread.daemon = True

        thread.start()
        trace.logger.debug("Conversion thread launched")

        return Response(response="{}", status=200, mimetype="application/json")

    except Exception as excpt:
        trace.logger.error(f"launch_conversion failed: {str(excpt)}")
        return Response(
            json.dumps({"error": str(excpt)}),
            status=500,
            mimetype="application/json"
        )


def conversion_thread(
    input_file_name,
    output_file_name,
    input_format,
    output_format,
    input_options,
    output_options,
    log_filename,
    sankey_as_data,
    process_label=None,
):
    """
    Thread de conversion universel.

    Parameters
    ----------
    input_file_name : str
        Chemin du fichier d'entrée
    output_file_name : str
        Chemin du fichier de sortie
    input_format : str
        Format d'entrée ('excel', 'json')
    output_format : str
        Format de sortie ('excel', 'json')
    input_options : dict
        Options de lecture du format d'entrée
    output_options : dict
        Options d'écriture du format de sortie
    trace_filename : str
        Fichier de trace utilisateur
    log_filename : str
        Fichier de logs debug
    process_label : str, optional
        Libellé localisé de l'opération, fourni par le dialogue appelant
        (« Chargement du fichier Excel », « Édition »…). Sert d'en-tête de
        bandeau à la place du générique « CONVERSION: EXCEL → JSON ». Si absent,
        on retombe sur le libellé technique.
    """
    trace.logger_init(log_filename, "a")
    write_process_status(log_filename, PROCESS_STATUS_RUNNING)

    # Le contrôle d'arrêt côté client repose désormais sur le fichier de statut,
    # plus sur ce texte : le bandeau peut donc être librement localisé/contextuel.
    banner_title = process_label or f"CONVERSION: {input_format.upper()} → {output_format.upper()}"
    op_label = process_label or "CONVERSION"

    trace.logger.info("=" * 80)
    trace.logger.info(banner_title)
    trace.logger.info(f"Input:  {Path(input_file_name).name}")
    trace.logger.info(f"Output: {Path(output_file_name).name}")
    trace.logger.debug(f"input_options: {input_options}")
    trace.logger.debug(f"output_options: {output_options}")
    trace.logger.info("=" * 80)

    t_total_start = perf_counter()

    try:
        # Ad-hoc shortcut: build the Index directly from the workbook's existing
        # sheets, without parsing/re-emitting the Sankey content. Sidesteps the
        # data-merge limitation of the full round-trip path (multiple
        # "Valeurs"-typed tabs are collapsed into one on regular write).
        if output_options.get("create_index_only"):
            if input_format != 'excel' or output_format != 'excel':
                raise ValueError(
                    "create_index_only requires excel input and output"
                )
            io_input = IOExcel()
            ok, msg = io_input.create_index_only(
                input_file_name,
                output_file_name,
                with_index=bool(output_options.get("with_index_sheet", True)),
                with_description=bool(
                    output_options.get("with_description_sheet", False)
                ),
                with_formatting=bool(
                    output_options.get("with_sheet_formating", False)
                ),
            )
            t_total = perf_counter() - t_total_start
            trace.logger.info("=" * 80)
            if ok:
                trace.logger.info(
                    f"✓ {op_label} en {t_total:.3f}s — Index créé"
                )
                write_process_status(log_filename, PROCESS_STATUS_FINISHED)
            else:
                trace.logger.error(
                    f"✗ {op_label} — échec après {t_total:.3f}s: {msg}"
                )
                write_process_status(log_filename, PROCESS_STATUS_FAILED)
            trace.logger.info("=" * 80)
            return

        # Ad-hoc shortcut : « Charger seulement la mise en page ». On ne parse
        # PAS le contenu Sankey (nœuds/données/TER) : on extrait uniquement
        # l'onglet caché « layout », qui contient déjà un JSON complet du
        # diagramme, et on le renvoie tel quel — exactement comme l'ouverture
        # d'un fichier JSON. Court-circuite load_sankey (donc réconciliation et
        # checks de cohérence), ce qui rouvre rapidement la mise en page d'un
        # Excel sans dépendre de la validité des onglets structurels.
        if input_options.get("only_layout"):
            if input_format != 'excel' or output_format != 'json':
                raise ValueError(
                    "only_layout requires excel input and json output"
                )
            with pd.ExcelFile(input_file_name) as excel_file:
                has_layout = "layout" in excel_file.sheet_names
            if not has_layout:
                t_total = perf_counter() - t_total_start
                trace.logger.error("=" * 80)
                trace.logger.error(
                    f"✗ {op_label} — échec après {t_total:.3f}s : aucun onglet "
                    "« layout » dans le fichier Excel (la mise en page n'est "
                    "présente que dans les fichiers exportés depuis l'application)"
                )
                trace.logger.error("=" * 80)
                write_process_status(log_filename, PROCESS_STATUS_FAILED)
                return
            # header=None : l'onglet « layout » est écrit en colonne A sans
            # en-tête (cf. write_sankey plus bas), une ligne par fragment.
            layout_table = pd.read_excel(input_file_name, "layout", header=None)
            layout_json_str = "".join(
                str(layout_table.iloc[_, 0]) for _ in range(len(layout_table)))
            # json.loads valide le contenu (échec propre si l'onglet est corrompu)
            layout_json = json.loads(layout_json_str)
            with open(output_file_name, "w", encoding="utf-8") as f:
                json.dump(layout_json, f)
            t_total = perf_counter() - t_total_start
            trace.logger.info("=" * 80)
            trace.logger.info(
                f"✓ {op_label} en {t_total:.3f}s — mise en page seule extraite "
                "(onglets structurels ignorés)"
            )
            write_process_status(log_filename, PROCESS_STATUS_FINISHED)
            trace.logger.info("=" * 80)
            return

        # Choisir le bon IO selon le format. 'blob' (= sankey courant) est
        # toujours matérialisé en .json sur disque avant d'arriver ici (cf.
        # launch_conversion / launch_optim), donc on le traite comme du json :
        # le state de session peut encore porter 'blob' quand retrieve_json
        # relance une conversion sur le fichier résultat.
        if input_format == 'excel':
            io_input = IOExcel()
        elif input_format in ('json', 'blob'):
            io_input = IOJson()
        else:
            raise ValueError(f"Format d'entrée '{input_format}' non supporté")

        # preserve_extra_columns est exposé dans l'onglet "Options de sortie"
        # côté UI (la décision est sémantiquement une décision d'écriture), mais
        # le stash des colonnes inconnues doit être armé pendant la lecture.
        # On propage donc le flag à input_options avant load_sankey.
        if "preserve_extra_columns" in output_options:
            input_options.setdefault(
                "preserve_extra_columns",
                output_options["preserve_extra_columns"],
            )

        # Active les coherence checks comme dans la réconciliation. Sans ça,
        # E6 (redondance de référence dans les contraintes) passe silencieusement
        # à la conversion et les toggles autofix_* du dialogue n'ont aucun effet
        # visible. Avec do_coherence_checks=True : E6 fait abort l'absence
        # d'autofix, et avec les toggles on les autofixes du load (E5/E6) tournent
        # avant le check, le silencent et accumulent les corrections sur
        # _auto_corrected_node_cells / _auto_corrected_constraint_ids.
        input_options.setdefault('do_coherence_checks', True)

        # Charger avec les options d'entrée
        trace.logger.info("📖 Lecture du fichier source...")
        t_read_start = perf_counter()
        ok, msg = io_input.load_sankey(input_file_name, **input_options)
        max_line_length = 50
        # input_options['layout'] (case « Onglet mise en page » du dialogue
        # d'ouverture) pilote la lecture de l'onglet caché « layout ». Décoché =>
        # on ignore la mise en page sauvegardée (positions/styles) pour laisser
        # le front recalculer une mise en page automatique. Défaut = True.
        if input_format == 'excel' and input_options.get('layout', True):
            try:
                # Vérifier que la sheet layout existe
                with pd.ExcelFile(input_file_name) as excel_file:
                    if "layout" in excel_file.sheet_names:
                        # header=None : l'onglet « layout » est écrit en colonne A
                        # lignes 1..N sans en-tête (cf. write ci-dessous). Sans ça,
                        # read_excel prend la 1re ligne comme nom de colonne et
                        # l'accès par label `[0]` lève KeyError sur pandas 2.x.
                        layout_table = pd.read_excel(input_file_name, "layout", header=None)
                        trace.logger.info("{:-<{w}}".format("Extract diagram layout ", w=max_line_length))
                        layout_json_str = "".join(
                            str(layout_table.iloc[_, 0]) for _ in range(len(layout_table)))
                        layout_json = json.loads(layout_json_str)

                        # Ajouter le layout aux options de sortie pour JSON
                        if output_format == 'json':
                            output_options['layout'] = layout_json
                            trace.logger.info("✓ Layout extracted and will be included in JSON")
                    else:
                        trace.logger.debug("No layout sheet found in Excel file")
            except Exception as e:
                trace.logger.warning(f"Could not extract layout: {e}")

        t_read = perf_counter() - t_read_start

        # Symmetric input-options contract: load_sankey returns ok=False only
        # when one of the six options is OFF and its problem was detected. That
        # is a deliberate abort — the user must enable the named option — so
        # the conversion hard-fails here. (No "best-effort" output: in strict
        # mode the offending node/flux is *detected* but not created, so the
        # Sankey would silently revert to its base structure.)
        if not ok:
            t_total = perf_counter() - t_total_start
            trace.logger.error("=" * 80)
            trace.logger.error(f"✗ {op_label} — échec après {t_total:.3f}s")
            for line in msg.split("\n"):
                if line.strip():
                    trace.logger.error(f"  {line}")
            trace.logger.error("=" * 80)
            write_process_status(log_filename, PROCESS_STATUS_FAILED)
            return

        # Load succeeded — inspect the autocorrect accumulators populated when
        # the create_new_* / autofix_* options were ON, to drive the red
        # highlighting in write_sankey below.
        added_nodes = getattr(io_input.sankey, "_auto_corrected_nodes", []) or []
        added_flux = getattr(io_input.sankey, "_auto_corrected_flux", []) or []
        node_cells = getattr(io_input.sankey, "_auto_corrected_node_cells", []) or []
        constraint_ids = getattr(io_input.sankey, "_auto_corrected_constraint_ids", []) or []
        autofix_applied = bool(added_nodes or added_flux or node_cells or constraint_ids)

        # Taille du fichier d'entrée
        input_size = Path(input_file_name).stat().st_size / (1024 * 1024)
        trace.logger.info(f"✓ Lecture terminée: {input_size:.2f} MB en {t_read:.3f}s")

        # Choisir le bon IO pour l'écriture
        if output_format == 'excel':
            io_output = IOExcel(io_input.sankey)
        elif output_format == 'json':
            io_output = IOJson(io_input.sankey)
        else:
            raise ValueError(f"Format de sortie '{output_format}' non supporté")

        # keep_other_sheets : converter-only flag — copie l'input Excel vers
        # l'output avant d'écrire pour préserver les onglets non-format.
        # Seul le `rewrite_format_sheets` est compris par write_sankey ; il
        # détermine ensuite si les onglets format présents sont remplacés ou
        # laissés intacts.
        keep_other_sheets = output_options.pop('keep_other_sheets', False)

        if output_format == 'excel' and keep_other_sheets:
            if input_format == 'excel':
                shutil.copyfile(input_file_name, output_file_name)
                trace.logger.info("✓ Input Excel copied to output to preserve other sheets")
            else:
                trace.logger.warning("keep_other_sheets ignored: input is not Excel")

        # Si la lecture tolérante a accumulé des corrections (toggles
        # create_new_* / autofix_* cochés), on log un résumé par catégorie et
        # on force highlight_autocorrect=True sur l'écriture Excel pour que les
        # cellules / lignes ajoutées ou corrigées apparaissent en rouge.
        if autofix_applied:
            if added_nodes:
                trace.logger.warning(
                    "Nœuds créés : {} nœud(s) absent(s) de l'onglet Noeuds, "
                    "créé(s) et surligné(s) en rouge (premiers : {})".format(
                        len(added_nodes), ", ".join('"{}"'.format(n) for n in added_nodes[:5])
                    )
                )
            if added_flux:
                trace.logger.warning(
                    "Flux créés / propagés : {} flux surligné(s) en rouge (premiers : {})".format(
                        len(added_flux),
                        ", ".join('"{}"->"{}"'.format(o, d) for o, d in added_flux[:5])
                    )
                )
            if node_cells:
                trace.logger.warning(
                    "Autofix mat_balance : {} cellule(s) corrigée(s) (premiers : {})".format(
                        len(node_cells), ", ".join('"{}"/{}'.format(n, c) for n, c in node_cells[:5])
                    )
                )
            if constraint_ids:
                trace.logger.warning(
                    "Autofix contraintes : {} id(s) avec doublons dédupliqués (premiers : {})".format(
                        len(constraint_ids), ", ".join(str(i) for i in constraint_ids[:5])
                    )
                )
            if output_format == 'excel':
                output_options['highlight_autocorrect'] = True

        # Écrire avec les options de sortie
        trace.logger.info("📝 Écriture du fichier de sortie...")
        t_write_start = perf_counter()
        io_output.write_sankey(output_file_name, **output_options)
        if input_format != 'excel':
            if "layout" in output_options and output_options["layout"]:
                # Ajoute le fichier json dans un onglet layout
                wb = openpyxl.load_workbook(output_file_name)
                layout_sheet = wb.create_sheet()
                layout_sheet.title = "layout"
                splitted_layout = cut_layout(sankey_as_data)
                cpt = 1
                for i in splitted_layout:
                    layout_sheet["A" + str(cpt)].value = i
                    cpt = cpt + 1
                # Masquer et protéger l'onglet layout
                layout_sheet.sheet_state = 'hidden'
                layout_sheet.protection.sheet = True
                wb.save(output_file_name)
        t_write = perf_counter() - t_write_start

        # Taille du fichier de sortie
        output_size = Path(output_file_name).stat().st_size / (1024 * 1024)
        trace.logger.info(f"✓ Écriture terminée: {output_size:.2f} MB en {t_write:.3f}s")

        # Temps total
        t_total = perf_counter() - t_total_start

        trace.logger.info("=" * 80)
        trace.logger.info(
            f"✓ {op_label} en {t_total:.3f}s "
            f"(lecture: {t_read:.3f}s, écriture: {t_write:.3f}s)"
        )
        trace.logger.info(
            f"  Taille: {input_size:.2f} MB → {output_size:.2f} MB "
            f"(ratio: {output_size/input_size:.2f}x)" if input_size > 0 else ""
        )
        trace.logger.info("=" * 80)
        write_process_status(log_filename, PROCESS_STATUS_FINISHED)

    except Exception as e:
        t_total = perf_counter() - t_total_start
        trace.logger.error("=" * 80)
        trace.logger.error(f"✗ {op_label} — échec après {t_total:.3f}s")
        trace.logger.error(f"Erreur: {str(e)}")
        write_process_status(log_filename, PROCESS_STATUS_FAILED)
        trace.logger.error("=" * 80)
        raise


@opensankey.route("/upload/clean", methods=["POST"])
def clean():
    set_process_state(process_started=False)
    return Response(response="{}", status=200, mimetype="application/json")


@opensankey.route("/example/download", methods=["POST"])
def download_examples():
    data_folder = os.environ.get("MFAData")
    # exemples_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exemples')
    exemple = request.get_data().decode("utf-8")
    exemple_file_path = os.path.join(data_folder, exemple)
    if os.path.exists(exemple_file_path):
        return send_file(exemple_file_path, as_attachment=True)
    return Response(exemple_file_path, status=400, mimetype="text")


@opensankey.route("/menus/excel_template", methods=["POST"])
def menus_excel_template():
    """Generate a blank Excel template with selected sheets, example rows, column notes and a Readme."""
    data = request.get_json(force=True)
    sheets = data.get("sheets", [])
    lang = data.get("lang", "fr")

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()

    try:
        _generate_excel_template(tmp.name, sheets, lang)
        return send_file(tmp.name, as_attachment=True, download_name="template_afm.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        trace.logger.error(f"Excel template generation failed: {e}")
        return Response(str(e), status=500)
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass


def _generate_excel_template(filepath, sheets, lang="fr"):
    """Build an openpyxl workbook with one tab per requested sheet."""
    from openpyxl.comments import Comment
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    example_font = Font(italic=True, color="888888")

    fr = lang == "fr"

    # ── Sheet definitions ──────────────────────────────────────────────
    SHEET_DEFS = {
        "tags": {
            "name": "Etiquettes" if fr else "Tags",
            "cols": [
                ("Nom du groupe d'étiquette" if fr else "Tags group name",
                 "Nom unique du groupe d'étiquettes." if fr else "Unique name for the tag group."),
                ("Type d'étiquette" if fr else "Tags type",
                 "Un parmi : dataTags, unitTags, fluxTags, nodeTags, levelTags" if fr
                 else "One of: dataTags, unitTags, fluxTags, nodeTags, levelTags"),
                ("Etiquettes" if fr else "Tags",
                 "Liste séparée par ':'" if fr else "Colon-separated list"),
                ("Palette visible" if fr else "Visible colormap",
                 "1 = affichée, 0 = masquée" if fr else "1 = shown, 0 = hidden"),
                ("Palette de couleur" if fr else "Colormap",
                 "Nom de palette (ex. Set1, Paired)" if fr else "Palette name (e.g. Set1, Paired)"),
                ("Couleurs" if fr else "Colors",
                 "Couleurs hex séparées par ':'" if fr else "Hex colors separated by ':'"),
            ],
            "examples": [
                ["Filière", "nodeTags", "Amont:Transformation:Aval", "1", "Set1", "#e41a1c:#377eb8:#4daf4a"],
                ["Source", "fluxTags", "Statistique:Estimation:Hypothèse", "1", "", ""],
                ["Année", "dataTags", "2023:2024", "", "", ""],
            ],
        },
        "nodes": {
            "name": "Noeuds" if fr else "Nodes",
            "cols": [
                ("Niveau d'agrégation" if fr else "Aggregation level",
                 "Entier >= 1. Niveau 1 = le plus agrégé." if fr
                 else "Integer >= 1. Level 1 = most aggregated."),
                ("Noeuds" if fr else "Nodes",
                 "Nom unique du noeud." if fr else "Unique node name."),
                ("Equilibre entrée-sortie" if fr else "Input-output balance",
                 "1 = bilan matière imposé, 0 = non." if fr
                 else "1 = mass balance enforced, 0 = no."),
            ],
            "examples": [
                [1, "Agriculture", 1],
                [2, "Céréales", 1],
                [2, "Élevage", 1],
            ],
        },
        "products": {
            "name": "Produits" if fr else "Products",
            "cols": [
                ("Niveau d'agrégation" if fr else "Aggregation level",
                 "Entier >= 1. Niveau 1 = le plus agrégé." if fr
                 else "Integer >= 1. Level 1 = most aggregated."),
                ("Produits" if fr else "Products",
                 "Nom unique du produit." if fr else "Unique product name."),
                ("Equilibre entrée-sortie" if fr else "Input-output balance",
                 "1 = bilan matière imposé, 0 = non." if fr
                 else "1 = mass balance enforced, 0 = no."),
            ],
            "examples": [
                [1, "Blé", 1],
                [1, "Lait", 1],
            ],
        },
        "sectors": {
            "name": "Secteurs" if fr else "Sectors",
            "cols": [
                ("Niveau d'agrégation" if fr else "Aggregation level",
                 "Entier >= 1." if fr else "Integer >= 1."),
                ("Secteurs" if fr else "Sectors",
                 "Nom unique du secteur." if fr else "Unique sector name."),
                ("Equilibre entrée-sortie" if fr else "Input-output balance",
                 "1 = bilan matière imposé, 0 = non." if fr
                 else "1 = mass balance enforced, 0 = no."),
            ],
            "examples": [
                [1, "Production", 1],
                [1, "Transformation", 1],
                [1, "Distribution", 1],
            ],
        },
        "exchanges": {
            "name": "Echanges" if fr else "Exchanges",
            "cols": [
                ("Niveau d'agrégation" if fr else "Aggregation level",
                 "Entier >= 1." if fr else "Integer >= 1."),
                ("Echanges" if fr else "Exchanges",
                 "Nom unique de l'échange." if fr else "Unique exchange name."),
            ],
            "examples": [
                [1, "Import"],
                [1, "Export"],
            ],
        },
        "ter": {
            "name": "Table emplois ressources" if fr else "Supply-use table",
            "cols": [
                ("" if fr else "",
                 "Produits en lignes, Secteurs en colonnes.\nTable Ressources (haut) puis Emplois (bas) séparées par une ligne vide.\n1 = flux existant, 0 ou vide = pas de flux."
                 if fr else "Products in rows, Sectors in columns.\nSupply table (top) then Use table (bottom) separated by blank row.\n1 = flow exists, 0 or empty = no flow."),
                ("Production", ""),
                ("Transformation", ""),
            ],
            "examples": [
                ["Blé", 1, ""],
                ["Lait", 1, ""],
                ["", "", ""],
                ["Blé", "", 1],
                ["Lait", "", 1],
            ],
        },
        "data": {
            "name": "Valeurs" if fr else "Values",
            "cols": [
                ("Origine" if fr else "Origin",
                 "Noeud d'origine du flux." if fr else "Flow origin node."),
                ("Destination" if fr else "Target",
                 "Noeud de destination du flux." if fr else "Flow destination node."),
                ("Valeur" if fr else "Value",
                 "Valeur dans l'unité de référence." if fr else "Value in reference unit."),
                ("Incertitude relative" if fr else "Relative Uncertainty",
                 "Écart-type relatif (ex. 0.1 = 10%)." if fr
                 else "Relative standard deviation (e.g. 0.1 = 10%)."),
                ("Source" if fr else "Source",
                 "Référence de la source." if fr else "Data source reference."),
            ],
            "examples": [
                ["Production", "Blé", 1500, 0.1, "FAO 2023"],
                ["Blé", "Transformation", 1200, 0.15, "Estimation"],
            ],
        },
        "min_max": {
            "name": "Min Max",
            "cols": [
                ("Origine" if fr else "Origin",
                 "Noeud d'origine du flux." if fr else "Flow origin node."),
                ("Destination" if fr else "Target",
                 "Noeud de destination du flux." if fr else "Flow destination node."),
                ("Minimum",
                 "Borne inférieure (unité de référence)." if fr else "Lower bound (reference unit)."),
                ("Maximum",
                 "Borne supérieure (unité de référence)." if fr else "Upper bound (reference unit)."),
            ],
            "examples": [
                ["Production", "Blé", 1000, 2000],
                ["Blé", "Transformation", 800, ""],
            ],
        },
        "constraints": {
            "name": "Contraintes" if fr else "Constraints",
            "cols": [
                ("ID",
                 "Identifiant de l'équation. Les lignes de même ID forment une équation." if fr
                 else "Equation ID. Rows with same ID form one equation."),
                ("Origine" if fr else "Origin",
                 "Noeud d'origine du flux." if fr else "Flow origin node."),
                ("Destination" if fr else "Target",
                 "Noeud de destination du flux." if fr else "Flow destination node."),
                ("eq = 0",
                 "Coefficient d'égalité : Σ(coef × flux) = 0" if fr
                 else "Equality coefficient: Σ(coef × flux) = 0"),
                ("eq <= 0",
                 "Coefficient d'inégalité : Σ(coef × flux) <= 0" if fr
                 else "Inequality coefficient: Σ(coef × flux) <= 0"),
                ("eq >= 0",
                 "Coefficient d'inégalité : Σ(coef × flux) >= 0" if fr
                 else "Inequality coefficient: Σ(coef × flux) >= 0"),
                ("Traduction" if fr else "Translation",
                 "Description lisible de la contrainte." if fr
                 else "Human-readable constraint description."),
            ],
            "examples": [
                [1, "Production", "Blé", -1, "", "", "Rendement 60%"],
                [1, "Céréales", "Production", 0.6, "", "", ""],
            ],
        },
        "ratio_flux": {
            "name": "Ratio Flux",
            "cols": [
                ("Origine" if fr else "Origin",
                 "Noeud d'origine du flux principal." if fr else "Main flow origin node."),
                ("Destination" if fr else "Destination",
                 "Noeud de destination du flux principal." if fr else "Main flow destination node."),
                ("=", "Colonne décorative." if fr else "Decorative column."),
                ("Coef",
                 "Coefficient d'égalité : flux = Coef × flux_ref" if fr
                 else "Equality coefficient: flow = Coef × ref_flow"),
                ("Min",
                 "Ratio minimum : flux >= Min × flux_ref" if fr
                 else "Minimum ratio: flow >= Min × ref_flow"),
                ("Max",
                 "Ratio maximum : flux <= Max × flux_ref" if fr
                 else "Maximum ratio: flow <= Max × ref_flow"),
                ("x", "Opérateur (décoratif)." if fr else "Operator (decorative)."),
                ("Origine Ref" if fr else "Origin Ref",
                 "Noeud d'origine du flux de référence. 'TOUT' = total entrant/sortant." if fr
                 else "Reference flow origin. 'TOUT'/'ALL' = total in/out."),
                ("Destination Ref" if fr else "Destination Ref",
                 "Noeud de destination du flux de référence." if fr
                 else "Reference flow destination."),
            ],
            "examples": [
                ["Production", "Blé", "=", 0.6, "", "", "x", "Céréales", "Production"],
                ["Blé", "Export", "=", "", 0.1, 0.3, "x", "Production", "Blé"],
            ],
        },
        "stocks": {
            "name": "Stocks",
            "cols": [
                ("Noeud" if fr else "Node",
                 "Nom du noeud concerné." if fr else "Node name."),
                ("Stock initial" if fr else "Initial stock",
                 "Valeur du stock initial." if fr else "Initial stock value."),
                ("Variation de stock" if fr else "Stock variation",
                 "Delta stock (Σentrants - Σsortants - Δstock = 0)." if fr
                 else "Stock variation (Σin - Σout - Δstock = 0)."),
                ("Incertitude" if fr else "Uncertainty",
                 "Incertitude relative." if fr else "Relative uncertainty."),
                ("Min", "Borne inférieure." if fr else "Lower bound."),
                ("Max", "Borne supérieure." if fr else "Upper bound."),
            ],
            "examples": [
                ["Transformation", "", 50, 0.2, 0, 100],
            ],
        },
        "results": {
            "name": "Résultats" if fr else "Results",
            "cols": [
                ("Origine" if fr else "Origin", ""),
                ("Destination" if fr else "Target", ""),
                ("Valeur réconciliée" if fr else "Reconciled value",
                 "Valeur calculée par le solveur." if fr else "Value computed by the solver."),
                ("Borne inférieure" if fr else "Lower boundary",
                 "Borne inf si variable libre." if fr else "Lower bound if free variable."),
                ("Borne supérieure" if fr else "Upper boundary",
                 "Borne sup si variable libre." if fr else "Upper bound if free variable."),
            ],
            "examples": [],
        },
        "analysis": {
            "name": "Analyses des résultats" if fr else "Results analysis",
            "cols": [
                ("Origine" if fr else "Origin", ""),
                ("Destination" if fr else "Target", ""),
                ("Valeur réconciliée" if fr else "Reconciled value", ""),
                ("Borne inférieure" if fr else "Lower boundary", ""),
                ("Borne supérieure" if fr else "Upper boundary", ""),
                ("Valeur non-réconciliée" if fr else "Unreconciled value",
                 "Valeur d'entrée." if fr else "Input value."),
                ("Incertitude relative non-réconciliée" if fr else "Unreconciled relative uncertainty",
                 "Écart-type relatif d'entrée (%)." if fr else "Input relative std dev (%)."),
                ("Nb sigmas" if fr else "Nb sigmas",
                 "Écart entrée/réconcilié en σ. <= 2 OK, > 3 à vérifier." if fr
                 else "Input/reconciled gap in σ. <= 2 OK, > 3 check."),
                ("Type de variable" if fr else "Variable type",
                 "mesuré, redondant, déterminable, libre" if fr
                 else "measured, redundant, determinable, free"),
            ],
            "examples": [],
        },
        "readme": {
            "name": "Readme",
            "cols": [],
            "examples": [],
            "is_readme": True,
        },
    }

    README_LINES = [
        "FORMAT EXCEL — ANALYSE DE FLUX DE MATIÈRE (AFM)",
        "",
        "Ce fichier est un modèle vierge pour construire une AFM." if fr
        else "This file is a blank template for building an MFA.",
        "",
        "STRUCTURE :",
        "- Étiquettes : groupes d'étiquettes (filtrage, couleur, agrégation)" if fr
        else "- Tags: tag groups (filtering, coloring, aggregation)",
        "- Noeuds / Produits / Secteurs / Échanges : définition des noeuds" if fr
        else "- Nodes / Products / Sectors / Exchanges: node definitions",
        "- Table emplois ressources (TER) : topologie matricielle" if fr
        else "- Supply-use table (TER): matrix topology",
        "- Valeurs : valeurs de flux, incertitudes" if fr
        else "- Values: flow values, uncertainties",
        "- Min Max : bornes des flux" if fr else "- Min Max: flow bounds",
        "- Contraintes : équations entre flux" if fr else "- Constraints: equations between flows",
        "- Ratio Flux : contraintes de ratio (format simplifié)" if fr
        else "- Ratio Flux: ratio constraints (simplified format)",
        "- Stocks : variations de stock par noeud" if fr
        else "- Stocks: stock variations per node",
        "",
        "CONVENTIONS :",
        "- ':' sépare les étiquettes dans les listes" if fr
        else "- ':' separates tags in lists",
        "- Cellule vide = toutes les étiquettes attribuées" if fr
        else "- Empty cell = all tags assigned",
        "- '0' = aucune étiquette (élément masqué)" if fr
        else "- '0' = no tag (element hidden)",
        "- Incertitude relative par défaut : 10%" if fr
        else "- Default relative uncertainty: 10%",
        "",
        "Documentation complète : FormatExcel.md" if fr
        else "Full documentation: FormatExcel.md",
    ]

    for sheet_key in sheets:
        defn = SHEET_DEFS.get(sheet_key)
        if defn is None:
            continue

        ws = wb.create_sheet(title=defn["name"])

        if defn.get("is_readme"):
            # Readme sheet: plain text lines
            ws.column_dimensions["A"].width = 100
            for i, line in enumerate(README_LINES, 1):
                cell = ws.cell(row=i, column=1, value=line)
                if i == 1:
                    cell.font = Font(bold=True, size=14)
                elif line.startswith("STRUCTURE") or line.startswith("CONVENTIONS"):
                    cell.font = Font(bold=True, size=12)
            continue

        # Header row
        for col_idx, (col_name, col_note) in enumerate(defn["cols"], 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            if col_note:
                cell.comment = Comment(col_note, "SankeyExcelParser")

        # Example rows
        for row_idx, row_data in enumerate(defn.get("examples", []), 2):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val if val != "" else None)
                cell.font = example_font
                cell.border = thin_border

        # Auto-width columns
        for col_idx, (col_name, _) in enumerate(defn["cols"], 1):
            max_len = len(str(col_name))
            for row_data in defn.get("examples", []):
                if col_idx - 1 < len(row_data):
                    max_len = max(max_len, len(str(row_data[col_idx - 1])))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 40)

    wb.save(filepath)


@opensankey.route("/menus/templates", methods=["POST"])
def menus_templates():
    """
    Renvoie l'index des modeles (templates).

    Les modeles migres vivent dans le submodule SankeyData (env SANKEY_DATA,
    sous-dossier templates/) ; repli sur MFAData/Modeles/Template/ tant que la
    migration n'est pas deployee partout. Les chemins file_path / img_path de
    l'index sont relatifs a la racine SANKEY_DATA (ex. templates/<diff>/...) et
    sont servis : data via /opensankey/convert/launch (example_json), images via
    /opensankey/menus/templates_asset/<path>.
    """
    sankey_data = os.environ.get("SANKEY_DATA")
    index_path = os.path.join(sankey_data, "templates", "index.json") if sankey_data else None
    if not (index_path and os.path.exists(index_path)):
        index_path = os.path.join(os.environ.get("MFAData"), "Modèles", "Template", "index.json")
    with open(index_path, encoding="utf-8") as file_index:
        data_index = json.load(file_index)
    response = Response(response=json.dumps(data_index), status=200, mimetype="application/json")
    return response


@opensankey.route("/menus/templates_asset/<path:asset>", methods=["GET"])
def menus_templates_asset(asset):
    """
    Sert un fichier (image de previsualisation, ...) depuis la racine SANKEY_DATA.

    Le chemin `asset` est relatif a SANKEY_DATA (ex. templates/essential/image/
    business_simple.png) ; send_from_directory neutralise les remontees de chemin.
    """
    sankey_data = os.environ.get("SANKEY_DATA")
    if not sankey_data:
        abort(404)
    # Route GET publique : on ne sert que le contenu des modeles (templates/).
    if not asset.replace("\\", "/").startswith("templates/"):
        abort(404)
    return send_from_directory(sankey_data, asset)


@opensankey.route("/menus/examples", methods=["POST"])
def menus_examples():
    """
    _summary_

    Returns
    -------
    :return: _description_
    :rtype: _type_
    """
    data_folder = os.environ.get("MFAData")
    menus = {}
    # try:
    parse_folder(data_folder, menus)
    context = {"exemples_menu": menus}
    json_data = json.dumps(context)
    response = Response(response=json_data, status=200, mimetype="application/json")
    # except Exception as expt:
    #     response = Response(
    #         response=str(expt),
    #         status=500,
    #         mimetype='application/json'
    #     )
    # Try to import images from MFAData/OpenSankey/image_preview to static/media
    try:
        current_folder = os.environ.get("MFAData")
        list_in_folder = os.listdir(current_folder)
        if "MFAData" in list_in_folder and "image_preview" in os.listdir(
            current_folder + "\\MFAData\\Formations\\Démos\\OpenSankey\\"
        ):
            folder_image = current_folder + "\\MFAData\\Formations\\Démos\\OpenSankey\\image_preview"
            for i in os.listdir(folder_image):
                if i not in os.listdir(image_template_folder):
                    os.symlink(folder_image + "\\" + i, image_template_folder + "\\" + i)
    except Exception as expt:
        print(str(expt))
        response = Response(response=str(expt), status=500, mimetype="application/json")
        return response

    return response


@opensankey.route("/menus/tutorials", methods=["POST"])
def menus_tutorials():
    """
    Liste les tutoriels disponibles dans SANKEY_DATA/tutorials.

    Renvoie une liste ordonnee [{file, title}] lue depuis tutorials/index.json
    si present ; sinon repli sur un listing des fichiers .json / .json.gz du
    dossier (titre = nom de fichier sans extension).
    """
    sankey_data = os.environ.get("SANKEY_DATA")
    folder = os.path.join(sankey_data, "tutorials") if sankey_data else None
    tutorials = []
    index_path = os.path.join(folder, "index.json") if folder else None
    if index_path and os.path.exists(index_path):
        with open(index_path, encoding="utf-8") as file_index:
            data_index = json.load(file_index)
        raw = data_index.get("tutorials", []) if isinstance(data_index, dict) else data_index
        for item in raw:
            if isinstance(item, dict) and item.get("file"):
                tutorials.append({"file": item["file"], "title": item.get("title", item["file"])})
    elif folder and os.path.isdir(folder):
        for name in sorted(os.listdir(folder)):
            if name == "index.json":
                continue
            if name.endswith(".json.gz") or name.endswith(".json"):
                title = name
                for ext in (".json.gz", ".json"):
                    if title.endswith(ext):
                        title = title[: -len(ext)]
                        break
                tutorials.append({"file": name, "title": title.replace("_", " ")})
    response = Response(response=json.dumps({"tutorials": tutorials}), status=200, mimetype="application/json")
    return response


@opensankey.route("/open_sankeymatic", methods=["POST"])
def open_sankeymatic():
    try:
        # Get input Excel filename
        text_input_file = request.files["file_content"]

        # Create conversion files
        tmp_dir = tempfile.mkdtemp()  # Tempory dir for conversion
        text_input_filename = os.path.join(tmp_dir, "toto.txt")
        text_input_file.save(text_input_filename)

        ok, msg, json_obj = sankeymatic.parse_sankeymatic_file(text_input_filename)
        if not ok:
            print(msg)
        clean_file(text_input_filename, "Clean_TXT")

        response = Response(response=json.dumps(json_obj), status=200, mimetype="application/json")
        return response
    except Exception as e:
        current_app.logger.error("OPEN SANKEY MATIC | {0}".format(e))
        abort(500)
    return Response(
        json.dumps({"output": "ERROR: load_process: le fichier tmp_log n'existe pas."}),
        status=500,
        mimetype="application/json",
    )


@opensankey.route("/url/load_json", methods=["POST"])
def url_load_json():
    """
    HTTP POST request to get file from url path
    Input : None
    Output : file content (raw, let browser handle decompression)
    """
    try:
        url_front = request.form["url"]
        # Requête HTTP pour récupérer le fichier
        response = requests.get(url_front)
        response.raise_for_status()

        # Retourner le contenu brut du fichier
        flask_response = make_response(response.content)

        # Important: NE PAS définir Content-Encoding: gzip
        # Laisser le navigateur gérer automatiquement la décompression
        if url_front.endswith(".gz"):
            flask_response.headers["Content-Type"] = "application/json"  # Type final attendu
        else:
            flask_response.headers["Content-Type"] = "application/json"

        return flask_response

    except Exception as e:
        print(f"Erreur : {e}")
        return {"error": str(e)}, 500


from . import views_export  # noqa
